import telebot
from telebot import types
from openpyxl import load_workbook, workbook
from openpyxl.styles import Alignment, PatternFill, Font
from os.path import join, abspath

btn_all = {'Start': 'Start', 'MENU': 'MENU',
           'Household chemicals': 'Household chemicals',
           'Icecream': 'Icecream',
           'Milk': 'Milk',
           'Body care': 'Body care'
           }

btn_H = {'1.': '1.', '2.': '2.', '3.': '3.', '4.': '4.', '5.': '5.',
         '6.': '6.', '7.': '7.', '8.': '8.', '9.': '9.', '10.': '10.',
         '11.': '11.', '12.': '12.', '13.': '13.', '14.': '14.', '15.': '15.',
         '16.': '16', '17.': '17.', '18.': '18.', '19.': '19.', '20.': '20.',
         '21.': '21.', '22.': '22.'
         }

btn_I = {'1)': '1)', '2)': '2)', '3)': '3)', '4)': '4)', '5)': '5)'
         }

btn_M = {'№1': '№1', '№2': '№2', '№3': '№3', '№4': '№4', '№5': '№5'
         }

btn_C = {'1~': '1~', '2~': '2~', '3~': '3~', '4~': '4~', '5~': '5~'
         }


class BotState:
    def __init__(self):
        self.bot_begin = 'Start'
        self.bot_menu = 'MENU'


bot = telebot.TeleBot('1137709917:AAEZNbC7R-qspaqqdJesoDmHthFHXO0v6qw')

data_path = join('.', "ListGreenWrapBot.xlsx")
data_path = abspath(data_path)
wb = load_workbook(filename=data_path, read_only=True)
wsn = list(wb.sheetnames)


@bot.message_handler(commands=['start'])
def key(message):
    key_start = types.InlineKeyboardMarkup()
    btn_start = types.InlineKeyboardButton(text="Start", callback_data="Start")
    key_start.add(btn_start)
    bot.send_message(message.chat.id, "Welcome "
                     + message.from_user.first_name
                     + ' ' + message.from_user.last_name + ", I am GreenWrapBot!", reply_markup=key_start)


@bot.message_handler(content_types=['text'])
def send_message(message):
    bot.send_message(message.chat.id, "Sorry, i do not understand you! Please write /start")


@bot.callback_query_handler(func=lambda mess: mess.data == 'Start')
def message(mess):
    key_mess = types.InlineKeyboardMarkup()
    for i in range(len(wsn)):
        btn_mess = types.InlineKeyboardButton(text=wsn[i], callback_data=wsn[i])
        key_mess.add(btn_mess)
    bot.send_message(mess.message.chat.id, "Let's get started. Select the appropriate category!", reply_markup=key_mess)


@bot.callback_query_handler(func=lambda mess: mess.data == btn_all['Household chemicals'])
def category_house(mess):
    #print(444)
    key_household = types.InlineKeyboardMarkup()
    ar = []
    text_col1 = []
    text_col2 = []
    wb.active = 0
    household = wb.active
    for row in household.iter_rows(min_row=2, max_row=None, min_col=1, max_col=1):
        for cell in row:
            btn_household = types.InlineKeyboardButton(text=cell.value, callback_data=cell.value)
            ar.append(btn_household)
            text_col1.append(cell.value)
    key_household.add(*ar)
    for col2 in household.iter_rows(min_row=2, max_row=None, min_col=2, max_col=2):
        for cell in col2:
            text_col2.append(cell.value)
    text1 = "\n".join(text_col2)
    bot.send_message(mess.message.chat.id, text="Find your good:" + "\n" + text1, reply_markup=key_household)


@bot.callback_query_handler(func=lambda mess: mess.data == btn_all['Icecream'])
def category_ice(mess):
    key_icecream = types.InlineKeyboardMarkup()
    ari = []
    text_col1i = []
    text_col2i = []
    wb.active = 1
    icecream = wb.active
    for row in icecream.iter_rows(min_row=2, max_row=None, min_col=1, max_col=1):
        for cell in row:
            btn_icecream = types.InlineKeyboardButton(text=cell.value, callback_data=cell.value)
            ari.append(btn_icecream)
            text_col1i.append(cell.value)
    key_icecream.add(*ari)
    for col2 in icecream.iter_rows(min_row=2, max_row=None, min_col=2, max_col=2):
        for cell in col2:
            text_col2i.append(cell.value)
    text1i = "\n".join(text_col2i)
    bot.send_message(mess.message.chat.id, text="Find your good:" + "\n" + text1i, reply_markup=key_icecream)


@bot.callback_query_handler(func=lambda mess: mess.data == btn_all['Milk'])
def category_milk(mess):
    key_milk = types.InlineKeyboardMarkup()
    arm = []
    text_col1m = []
    text_col2m = []
    wb.active = 2
    milk = wb.active
    for row in milk.iter_rows(min_row=2, max_row=None, min_col=1, max_col=1):
        for cell in row:
            btn_milk = types.InlineKeyboardButton(text=cell.value, callback_data=cell.value)
            arm.append(btn_milk)
            text_col1m.append(cell.value)
    key_milk.add(*arm)
    for col2 in milk.iter_rows(min_row=2, max_row=None, min_col=2, max_col=2):
        for cell in col2:
            text_col2m.append(cell.value)
    text1m = "\n".join(text_col2m)
    bot.send_message(mess.message.chat.id, text="Find your good:" + "\n" + text1m, reply_markup=key_milk)


@bot.callback_query_handler(func=lambda mess: mess.data == btn_all['Body care'])
def category_care(mess):
    key_care = types.InlineKeyboardMarkup()
    arc = []
    text_col1c = []
    text_col2c = []
    wb.active = 3
    care = wb.active
    for row in care.iter_rows(min_row=2, max_row=None, min_col=1, max_col=1):
        for cell in row:
            btn_care = types.InlineKeyboardButton(text=cell.value, callback_data=cell.value)
            arc.append(btn_care)
            text_col1c.append(cell.value)
    key_care.add(*arc)
    for col2 in care.iter_rows(min_row=2, max_row=None, min_col=2, max_col=2):
        for cell in col2:
            text_col2c.append(cell.value)
    text1c = "\n".join(text_col2c)
    bot.send_message(mess.message.chat.id, text="Find your good:" + "\n" + text1c, reply_markup=key_care)


@bot.callback_query_handler(func=lambda mess: mess.data in btn_H)
def good_household(mess):
    key_household = types.InlineKeyboardMarkup()
    # btn_household =
    wb.active = 0
    mark = str()
    household = wb.active
    for row in household.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
        if mess.data[:len(mess.data) - 1] == row[1][:len(mess.data) - 1]:
            mark = row[2]
        else:
            print('-1')
    bot.send_message(mess.message.chat.id, text="Marking:   " + mark, reply_markup=key_household)


@bot.callback_query_handler(func=lambda mess: mess.data in btn_I)
def good_icecream(mess):
    key_icecream = types.InlineKeyboardMarkup()
    # btn_icecream =
    marki = str()
    wb.active = 1
    icecream = wb.active
    for row in icecream.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
        if mess.data == row[1][:2]:
            marki = row[2]
        else:
            print('-1')
    bot.send_message(mess.message.chat.id, text="Marking:   " + marki, reply_markup=key_icecream)


@bot.callback_query_handler(func=lambda mess: mess.data in btn_M)
def good_milk(mess):
    key_milk = types.InlineKeyboardMarkup()
    # btn_milk =
    markm = str()
    wb.active = 2
    milk = wb.active
    for row in milk.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
        if mess.data[1] == row[1][0]:
            markm = row[2]
        else:
            print('-1')
    bot.send_message(mess.message.chat.id, text="Marking:   " + markm, reply_markup=key_milk)


@bot.callback_query_handler(func=lambda mess: mess.data in btn_C)
def good_care(mess):
    key_care = types.InlineKeyboardMarkup()
    # btn_care =
    markc = str()
    wb.active = 3
    care = wb.active
    for row in care.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
        if mess.data[:len(mess.data) - 1] == row[1][:len(mess.data) - 1]:
            markc = row[2]
        else:
            print('-1')
    bot.send_message(mess.message.chat.id, text="Marking:   " + markc, reply_markup=key_care)


bot.polling()
